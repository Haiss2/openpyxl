import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class AvgEachTypeOfPost():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        avg_each_type_post_list = wb.create_sheet("Avg of each type post")
        avg_each_type_post_list.sheet_view.zoomScale = 85
        
        for x in list(string.ascii_uppercase[0:9]):
            avg_each_type_post_list.column_dimensions[x].width = 14
        
        for merge_cells in ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:H1', 'I1:I2']:
            avg_each_type_post_list.merge_cells(merge_cells)

        data = [{
                "type": "Carousel",
                "number_of_posts": 10,
                "reach": 445,
                "impression": 500,
                "like": 23,
                "share": 45,
                "comment": 34,
                "total": 55,
                "engagement_rate": 7.48,
            }
            for i in range(3)]

        for y in range(3, 3 + len(data)):
            avg_each_type_post_list.row_dimensions[y].height = 40
        avg_each_type_post_list.row_dimensions[1].height = 25
        avg_each_type_post_list.row_dimensions[2].height = 25


        for row in avg_each_type_post_list['A1:I{}'.format(len(data)+2)]:
            for cell in row:
                cell.border = self.border

        headers = [
            { "label": "Type", "pos": "A1"},
            { "label": "Number of posts", "pos": "B1"},
            { "label": "Reach", "pos": "C1"},
            { "label": "Impression", "pos": "D1"},
            { "label": "Engagement", "pos": "E1"},
            { "label": "Likes", "pos": "E2"},
            { "label": "Comments", "pos": "F2"},
            { "label": "Shares", "pos": "G2"},
            { "label": "Total", "pos": "H2"},
            { "label": "Engagement Rate", "pos": "I1"},
            
        ]

        for header in headers:
            cell = avg_each_type_post_list[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header["label"]
        
        for i, row in enumerate(data):
            avg_each_type_post_list['A{}'.format(3+i)].value = row["type"]
            avg_each_type_post_list['B{}'.format(3+i)].value = row["number_of_posts"]
            avg_each_type_post_list['C{}'.format(3+i)].value = row["reach"]
            avg_each_type_post_list['D{}'.format(3+i)].value = row["impression"]
            avg_each_type_post_list['E{}'.format(3+i)].value = row["like"]
            avg_each_type_post_list['F{}'.format(3+i)].value = row["comment"]
            avg_each_type_post_list['G{}'.format(3+i)].value = row["share"]
            avg_each_type_post_list['H{}'.format(3+i)].value = row["total"]
            avg_each_type_post_list['I{}'.format(3+i)].value = row["engagement_rate"]

            for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                avg_each_type_post_list['{}{}'.format(letter, 3+i)].alignment = Alignment(vertical="center", horizontal="right")


         # Notions
        note_row = len(data) + 4
        avg_each_type_post_list['G' + str(note_row)].value = "※1 Stories are not included."

        avg_each_type_post_list.merge_cells('A{}:A{}'.format(len(data) + 6, len(data) + 8))
        cmt = avg_each_type_post_list['A{}'.format(len(data) + 6)]
        cmt.value = 'Comments'
        avg_each_type_post_list.merge_cells('B{}:I{}'.format(len(data) + 6, len(data) + 8))
        self.set_border(avg_each_type_post_list, 'B{}:I{}'.format(len(data) + 6, len(data) + 8))
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="003366FF")
