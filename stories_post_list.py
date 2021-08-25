import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class StoriesPostList():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        stories_post_list = wb.create_sheet("Stories Posts List")
        stories_post_list.sheet_view.zoomScale = 85
        
        stories_post_list.column_dimensions['A'].width = 24
        stories_post_list.column_dimensions['B'].width = 10
        for character in list(string.ascii_uppercase[2:9]):
            stories_post_list.column_dimensions[character].width = 20
        
        for character in list(string.ascii_uppercase[0:9]):
            stories_post_list.merge_cells('{}1:{}2'.format(character, character))

        data = [{
                "time": "2021/05/04 18:44:09",
                "photo": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQlu-5sicFuHii8BAVf-lwWzS0D4bOJ00mHAQ&usqp=CAU",
                "post_contents": "Styles are used to change the look of your data while displayed on screen. They are also used to determine the formatting for numbers.",
                "reach": 445,
                "reach_rate": 12,
                "impression": 500,
                "reply": 23,
                "move_from_stories": 45,
                "number_of_tap_on_next": 34,
                "number_of_tap_on_previos": 10,
            }
            for i in range(20)]

        for y in range(3, 3 + len(data)):
            stories_post_list.row_dimensions[y].height = 88
        stories_post_list.row_dimensions[1].height = 30
        stories_post_list.row_dimensions[2].height = 30


        for row in stories_post_list['A1:I{}'.format(len(data)+3)]:
            for cell in row:
                cell.border = self.border

        headers = ["Time", "Photo", 'Reach', "Reach Rate", "Impression", "Reply", "Move from Stories", "Number of Tap on 'Next'", "Number of Tap on 'Previos'"]

        for i, header in enumerate(headers):
            cell = stories_post_list[string.ascii_uppercase[i] + '1']
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header
        
        for i, row in enumerate(data):
            stories_post_list['A{}'.format(3+i)].value = row["time"]
            stories_post_list['C{}'.format(3+i)].value = row["reach"]
            stories_post_list['D{}'.format(3+i)].value = row["reach_rate"]
            stories_post_list['E{}'.format(3+i)].value = row["impression"]
            stories_post_list['F{}'.format(3+i)].value = row["reply"]
            stories_post_list['G{}'.format(3+i)].value = row["move_from_stories"]
            stories_post_list['H{}'.format(3+i)].value = row["number_of_tap_on_next"]
            stories_post_list['I{}'.format(3+i)].value = row["number_of_tap_on_previos"]
            for letter in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                stories_post_list['{}{}'.format(letter, 3+i)].alignment = Alignment(vertical="center")

            img = Image("img/stories.png")
            stories_post_list.add_image(img, 'B{}'.format(3+i))
        
        avg_data = [662, 749, 45, 0, 20, 65, 10]
        avg_row = len(data) + 3
        stories_post_list.merge_cells('A{}:B{}'.format(avg_row, avg_row)) 
        el = stories_post_list['A' + str(avg_row)]
        el.value  = 'Average'
        el.font  = Font(b=True, size=12)
        el.fill = PatternFill("solid", fgColor="00969696")
        el.alignment = Alignment(horizontal="right", vertical="center")
        for i, datum in enumerate(list(string.ascii_uppercase[2:9])):
            el = stories_post_list[datum + str(avg_row)]
            el.value = avg_data[i]
            el.font  = Font(b=True, size=12)
        

        # Notions
        note_row = len(data) + 5
        stories_post_list['E' + str(note_row)].value = "※1 The data of Stories can only be avaibale after the day your acc is connected to Reposta.ストーリーズはRepostaでアカウント連携した日以降のデータのみが表示されます。"
        stories_post_list['E' + str(note_row + 1)].value = "※2 The data via Highlight is not reflected on this chart. ハイライト経由での数値は反映されません。"
        stories_post_list['E' + str(note_row + 2)].value = "※3 In case you share the other user's post on your Stories, only the post of Buisiness account can be displayed. 他のユーザーの投稿をストーリーズでシェアした場合、ビジネスアカウントの投稿のみが表示されます。"

        stories_post_list['L' + str(avg_row)].alignment = Alignment(horizontal="right", vertical="center")

        stories_post_list.merge_cells('A{}:A{}'.format(len(data) + 9, len(data) + 11))
        cmt = stories_post_list['A{}'.format(len(data) + 9)]
        cmt.value = 'Comments'
        stories_post_list.merge_cells('B{}:I{}'.format(len(data) + 9, len(data) + 11))
        self.set_border(stories_post_list, 'B{}:I{}'.format(len(data) + 9, len(data) + 11))
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="003366FF")

