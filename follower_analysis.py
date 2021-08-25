import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart

class FollowerAnalysis():

    def __init__(self, wb):
        self.wb = wb

    def render(self):
        wb = self.wb

        # Follower Analysis
        follower_analysis = wb.create_sheet("Follower Analysis")
        follower_analysis.sheet_view.showGridLines = False

        follower_analysis.column_dimensions['A'].width = 32
        a_to_z = list(string.ascii_uppercase)
        for letter in a_to_z[1: 11]:
            follower_analysis.column_dimensions[letter].width = 12


        age_labels = ['13-17 yrs old', '18-24 yrs old', '25-34 yrs old', '35-44 yrs old', '45-54 yrs old', '55-64 yrs old', '+65 yrs old']
        age_data = [12,24,56,32,23,42,12]

        gender_labels = ['Male', 'Female']
        gender_data = [35, 65]

        country_labels = ['Japan' ,'Russia' ,'Iraq' ,'Ukrain' ,'USA' ,'Italy' ,'Brazil' ,'Israel' ,'Morocco' ,'India']
        country_data = [97,23,4,1,2,4,5,6,7,8]

        city_labels = ['Saitama-shi', 'Yokohama', 'Osaka', 'Nagoya-shi', 'Sapporo-shi', 'Kyoto-shi', 'Fukuoka-shi', 'Setagaya-ku', 'Sendai', 'Chiba-shi']
        city_data = [97,23,4,1,2,4,5,6,7,8]

        usage_rate_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        usage_rate_data = [67, 88, 89, 92, 87, 82, 92]

        active_follower_labels = ['0時~2時', '3時~5時', '6時~8時', '9時~11時', '12時~14時', '15時~17時', '18時~20時', '21時~23時']
        active_follower_data = [195, 	112, 	350, 	384, 	439, 	450, 	523, 	517] 


        for merge_cells in ['A1:A2', 'A3:A4', 'A5:A6', 'A7:A8', 'A9:A10', 'A11:A12']:
            follower_analysis.merge_cells(merge_cells)

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # border
        for row in follower_analysis['A1:K12']:
            for cell in row:
                cell.border = border
        for row in follower_analysis['B43:K43']:
            for cell in row:
                cell.border = border

        # blue title
        for x, y in [('Age', 'A1'), ('Male vs female', 'A3'), ('Country', 'A5'), ('City', 'A7'), ('Instagram usage rate by follower\n(based on date)', 'A9'), ('Active Follower Number\n(Based on Time)', 'A11'), ('Comment', 'A43')]:
            element = follower_analysis[y]
            element.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            element.value = x
            element.font  = Font(b=True, color="00FFFFFF", size=12)
            element.fill = PatternFill("solid", fgColor="003366FF")
            
        
        # fill color header
        for i in range(10):
            for j in range(6):
                element = follower_analysis.cell(row=2*j+1, column=i+2)
                element.fill = PatternFill("solid", fgColor="00C0C0C0")
                element.font  = Font(b=True, color="00000000", size=11)

        def render_data_row(labels, data, row):
            for c, col in enumerate(labels):
                title = follower_analysis.cell(row=row, column=c+2)
                title.value = labels[c]

                val = follower_analysis.cell(row=row+1, column=c+2)
                val.value = data[c]/100
                val.number_format = '0%'
        
        render_data_row(age_labels, age_data, 1)
        render_data_row(gender_labels, gender_data, 3)
        render_data_row(country_labels, country_data, 5)
        render_data_row(city_labels, city_data, 7)
        render_data_row(usage_rate_labels, usage_rate_data, 9)
        render_data_row(active_follower_labels, active_follower_data, 11)

        def render_pie_chart(row_data, title, where, max_col):
            pie = PieChart()
            pie.width = 8
            pie.height = 6
            labels = Reference(follower_analysis, min_row=row_data,min_col=2,max_col=max_col)
            data = Reference(follower_analysis, min_row=row_data+1,min_col=1,max_col=max_col)

            pie.add_data(data, titles_from_data=True, from_rows=True)
            pie.set_categories(labels)
            pie.title = title

            follower_analysis.add_chart(pie, where)
        
        render_pie_chart(1, 'Age', 'A14', 9)
        render_pie_chart(3, 'Male vs Female', 'D14', 4)
        render_pie_chart(5, 'Country', 'A26', 12)
        render_pie_chart(7, 'City', 'D26', 12)

        def render_bar_chart(row_data, title, where, max_col):
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.top = 50
            chart.left = 50
            chart.width = 8
            chart.height = 6
            chart.title = title

            labels = Reference(follower_analysis, min_row=row_data,min_col=2,max_col=max_col)
            data = Reference(follower_analysis, min_row=row_data+1,min_col=2,max_col=max_col, max_row=row_data+2)
            chart.add_data(data, titles_from_data=True, from_rows=True)
            chart.set_categories(labels)
            chart.shape = 4
            follower_analysis.add_chart(chart, where)


        render_bar_chart(9, 'フォロワーの\nInstagramアプリ利用率（曜日別）', 'H14', 8)
        render_bar_chart(11, 'Active Follower Number (Based on Time)', 'H26', 9)

        for merge_cells in ['A40:K40', 'A41:K41', 'B43:K43']:
            follower_analysis.merge_cells(merge_cells)
        
        notion1  = follower_analysis['A40']
        notion1.value = '※１"Instagram usage rate by follower (based on date)" shows on which date the follower use IG app most. 100% means the average line.'
        notion1.alignment = Alignment(horizontal="right", vertical="center")

        notion2 = follower_analysis['A41']
        notion2.value = '※２「Active Follower Number (Based on Time)」shows followers who were active during that time.'
        notion2.alignment = Alignment(horizontal="right", vertical="center")

        follower_analysis.row_dimensions[43].height = 40
        comment = follower_analysis['B43']
        comment.fill = PatternFill("solid", fgColor="00FFFFCC")
        comment.alignment = Alignment(vertical="center")



