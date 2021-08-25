import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class MoM():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        self.col_labes = [*list(string.ascii_uppercase), 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)

    def render_grid(self, ws, possition, img, unit_of_measure, title, data, pos_x, pos_y):
        self.set_border(ws, possition)
        img = Image(img)
        ws.add_image(img, self.col_labes[self.col_labes.index(pos_x) + 3] + str(pos_y + 1))

        ws.merge_cells('{}{}:{}{}'.format(self.col_labes[self.col_labes.index(pos_x) + 4], pos_y + 2, self.col_labes[self.col_labes.index(pos_x) + 5], pos_y + 3))
        number = ws['{}{}'.format(self.col_labes[self.col_labes.index(pos_x) + 4], pos_y + 2)]
        number.value = 10

        number.font = Font(color="00000000", size=24)
        number.alignment = Alignment(vertical="center")

        ws.merge_cells('{}{}:{}{}'.format(self.col_labes[self.col_labes.index(pos_x) + 6], pos_y + 2, self.col_labes[self.col_labes.index(pos_x) + 6], pos_y + 3))
        unit = ws['{}{}'.format(self.col_labes[self.col_labes.index(pos_x) + 6], pos_y + 2)]
        unit.value = unit_of_measure
        unit.alignment = Alignment(vertical="center")
        
        ws.merge_cells('{}{}:{}{}'.format(pos_x, pos_y + 4, self.col_labes[self.col_labes.index(pos_x) + 6], pos_y + 4))
        title_cell = ws['{}{}'.format(pos_x, pos_y + 4)]
        title_cell.value = title
        title_cell.font = Font(color="00969696", size=20)
        title_cell.alignment = Alignment(horizontal="center")

        # border
        for row in ws['{}{}:{}{}'.format(pos_x, pos_y+6, self.col_labes[self.col_labes.index(pos_x) + 6], pos_y+8)]:
            for cell in row:
                cell.border = self.border
        
        labels = ['2月', '3月', '4月', '5月', '6月', '7月', '8月']

        for c, col in enumerate(labels):
            header = ws.cell(row=pos_y+6, column=self.col_labes.index(pos_x) + 1 + c)
            header.value = col
            header.fill = PatternFill("solid", fgColor="00969696")

            val = ws.cell(row=pos_y+7, column=self.col_labes.index(pos_x) + 1 + c)
            val.value = data[c]
            # val.number_format = '0%'

            diff = ws.cell(row=pos_y+8, column=self.col_labes.index(pos_x) + 1 + c)
            diff.value = '-'
            diff.fill = PatternFill("solid", fgColor="00CCCCFF")

        # chart here
        c1 = LineChart()
        c1.width = 12.9
        c1.height = 7
        c1.style = 13

        labels = Reference(ws, min_row=pos_y+6, min_col=self.col_labes.index(pos_x) + 1,max_col=self.col_labes.index(pos_x) + 7)
        data = Reference(ws, min_row=pos_y+7, min_col=self.col_labes.index(pos_x), max_col=self.col_labes.index(pos_x) + 7)

        c1.add_data(data, titles_from_data=True, from_rows=True)
        c1.set_categories(labels)

        ws.add_chart(c1, "{}{}".format(pos_x, pos_y+9))

    def render(self):
        wb = self.wb
        m_o_m = wb.create_sheet("MoM") 
        m_o_m.sheet_view.showGridLines = False
        m_o_m.sheet_view.zoomScale = 60

        
        for x in self.col_labes:
            m_o_m.column_dimensions[x].width = 10
        for y in range(1, 60):
            m_o_m.row_dimensions[y].height = 25

        # render_grid(ws, possition, img, unit_of_measure, title, data, pos_x, pos_y):
        val1 = [8, 7, 9, 4, 3, 2, 10]
        self.render_grid(m_o_m, 'B3:H19', 'img/users.png', 'people', 'Number of Followers', val1, 'B', 3)
        self.render_grid(m_o_m, 'J3:P19', 'img/posts.png', 'posts', 'Number of Posts', val1, 'J', 3)
        self.render_grid(m_o_m, 'R3:X19', 'img/laptop.png', 'clicks', 'Number of Clicks on Website', val1, 'R', 3)
        self.render_grid(m_o_m, 'Z3:AF19', 'img/user.png', 'times', 'Number of Profile View', val1, 'Z', 3)
        self.render_grid(m_o_m, 'B23:H39', 'img/hand_shake.png', '%', 'Engagement Rate', val1, 'B', 23)
        self.render_grid(m_o_m, 'J23:P39', 'img/heart.png', 'likes', 'Number of Likes', val1, 'J', 23)
        self.render_grid(m_o_m, 'R23:X39', 'img/comment.png', 'cmts', 'Number of Comments', val1, 'R', 23)
        self.render_grid(m_o_m, 'Z23:AF39', 'img/download.png', 'times', 'Number of Saves', val1, 'Z', 23)
        self.render_grid(m_o_m, 'B43:H59', 'img/search_people.png', 'people', 'Number of Reaches', val1, 'B', 43)
        self.render_grid(m_o_m, 'J43:P59', 'img/eyes.png', 'times', 'Number of Impression', val1, 'J', 43)
            
        
        m_o_m['R44'].value = "※1"
        m_o_m['S44'].value = 'The figure above is monthly basis total number.'
        m_o_m['R45'].value = "※2"
        m_o_m['S45'].value = "The Number of Posts doesn't include Stories."
        m_o_m['R46'].value = "※3"
        m_o_m['S46'].value = "The figure of Number of Fellowers can only be shown since the day your acc is connected to Reposta."
        m_o_m['R47'].value = "※4"
        m_o_m['S47'].value = "The Number of Followers is the figure at the end of each month."
        m_o_m['R48'].value = "※5"
        m_o_m['S48'].value = "It doesn't include the figure of advertisement."
        m_o_m['R49'].value = "※6"
        m_o_m['S49'].value = "Engagemen Rate = Total Engagement number of all posts of the monh / Number of Reaches"
            
        
        for merge_cells in ['R51:AF52', 'R53:AF59', 'AC43:AC44', 'AD43:AF44']:
            m_o_m.merge_cells(merge_cells)
        
        m_o_m['AC43'].value = 'Example'
        m_o_m['AC43'].alignment = Alignment(horizontal="center", vertical="center")

        m_o_m['AD43'].value = "Comparing to\nPrevious Months"
        m_o_m['AD43'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.set_border(m_o_m, 'AD43:AF44')
        m_o_m['AD43'].fill = PatternFill("solid", fgColor="00CCCCFF")

        
        m_o_m['R51'].value = 'Comments'
        self.set_border(m_o_m, 'R53:AF59')
        m_o_m['R51'].alignment = Alignment(horizontal="center", vertical="center")
        m_o_m['R51'].font  = Font(b=True, color="00FFFFFF", size=12)
        m_o_m['R51'].fill = PatternFill("solid", fgColor="003366FF")
