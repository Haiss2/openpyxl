import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class WordDescription():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def render(self):
        wb = self.wb
        word_des = wb.create_sheet("Word Description") 
        word_des.sheet_view.zoomScale = 120
        
        for row in word_des['A1:B13']:
            for cell in row:
                cell.border = self.border

        word_des.column_dimensions['A'].width = 25
        word_des.column_dimensions['B'].width = 65

    
        word_des['A1'].value = "Word"
        word_des['A1'].fill = PatternFill("solid", fgColor="003366FF")
        word_des['A1'].font  = Font(b=True, color="00FFFFFF")
        word_des['A1'].alignment = Alignment(horizontal="center", vertical="center")

        word_des['B1'].value = "Description"
        word_des['B1'].fill = PatternFill("solid", fgColor="003366FF")
        word_des['B1'].font  = Font(b=True, color="00FFFFFF")
        word_des['B1'].alignment = Alignment(horizontal="center", vertical="center")

        data = [
            ("Number of Follower", "フォローされているアカウント数"),
            ("Follower (+/-)", "Number of Follower Changes compared to the day before"),
            ("Number of Website clicks", "Number of clicks on Profile URL"),
            ("Profile View", "Number of time the profile is viewed"),
            ("Reach", "Unique Number of the post being displayed"),
            ("Reach Rate", "Reach Number / Folower Number"),
            ("Impression", "Total Number of times that the post is displayed"),
            ("Like!", "Numbeer of Like! On the post"),
            ("Comment", "Numbeer of Comments On the post"),
            ("Save", "Number of User account who saved the post"),
            ("Engagement", "Rate	Engagement number / Reach Number"),
            ("Carousel", "Post that contains multiple photos & movies at the same time"),
        ] 
        
        for i, datum in enumerate(data):
            word_des['A' + str(i+2)].value = datum[0]
            word_des['B' + str(i+2)].value = datum[1]












