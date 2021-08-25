import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class TopWorstStory():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        top_worst_story = wb.create_sheet("Stories - Impression TOP&WORST3")
        top_worst_story.sheet_view.zoomScale = 85
        
        top_worst_story.column_dimensions['A'].width = 14
        top_worst_story.column_dimensions['B'].width = 24
        top_worst_story.column_dimensions['C'].width = 10
        for character in list(string.ascii_uppercase[3:9]):
            top_worst_story.column_dimensions[character].width = 20
        
        for character in list(string.ascii_uppercase[0:9]):
            top_worst_story.merge_cells('{}1:{}2'.format(character, character))

        datum = {
                "time": "2021/05/04 18:44:09",
                "photo": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQlu-5sicFuHii8BAVf-lwWzS0D4bOJ00mHAQ&usqp=CAU",
                "post_contents": "Styles are used to change the look of your data while displayed on screen. They are also used to determine the formatting for numbers.",
                "reply": 23,
                "move_from_stories": 45,
                "number_of_tap_on_next": 34,
                "number_of_tap_on_previos": 10,
                "reach": 20,
                "impression": 30
            }
        
        top_data = [datum for i in range(3)]
        worst_data = [datum for i in range(3)]
        


        headers = ["", "Posted Date", "Photo", "Reply", "Move from Stories", "Number of Tap on 'Next'", "Number of Tap on 'Previos'", 'Reach', 'Impression']


        for i, header in enumerate(headers):
            cell = top_worst_story[string.ascii_uppercase[i] + '1']
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header

        
        top_worst_story.row_dimensions[1].height = 30
        top_worst_story.row_dimensions[2].height = 30
        for y in [4,5,6,8,9,10]:
            top_worst_story.row_dimensions[y].height = 88


        for row in top_worst_story['A1:I10']:
            for cell in row:
                cell.border = self.border


        # Top of the world
        top_worst_story['A3'].value = 'TOP▲'
        top_worst_story['A3'].font = Font(b=True, color="00000000", size=12)
        for row in top_worst_story['A3:I3']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="0099CC00")

        for i, row in enumerate(top_data):
            no = top_worst_story['A{}'.format(4+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00000000", size=16)
            no.fill = PatternFill("solid", fgColor="0099CC00")

            top_worst_story['B{}'.format(4+i)].value = row["time"]
            top_worst_story['D{}'.format(4+i)].value = row["reply"]
            top_worst_story['E{}'.format(4+i)].value = row["move_from_stories"]
            top_worst_story['F{}'.format(4+i)].value = row["number_of_tap_on_next"]
            top_worst_story['G{}'.format(4+i)].value = row["number_of_tap_on_previos"]
            top_worst_story['H{}'.format(4+i)].value = row["reach"]
            top_worst_story['I{}'.format(4+i)].value = row["impression"]
            for letter in ['A', 'B' 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                top_worst_story['{}{}'.format(letter, 4+i)].alignment = Alignment(vertical="center")

            img = Image("img/stories.png")
            top_worst_story.add_image(img, 'C{}'.format(4+i))

        # Worst of the world
        top_worst_story['A7'].value = 'WORST▼'
        top_worst_story['A7'].font = Font(b=True, color="00FFFFFF", size=12)
        for row in top_worst_story['A7:I7']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="00FF6600")

        for i, row in enumerate(worst_data):
            no = top_worst_story['A{}'.format(8+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00FFFFFF", size=16)
            no.fill = PatternFill("solid", fgColor="00FF6600")

            top_worst_story['B{}'.format(8+i)].value = row["time"]
            top_worst_story['D{}'.format(8+i)].value = row["reply"]
            top_worst_story['E{}'.format(8+i)].value = row["move_from_stories"]
            top_worst_story['F{}'.format(8+i)].value = row["number_of_tap_on_next"]
            top_worst_story['G{}'.format(8+i)].value = row["number_of_tap_on_previos"]
            top_worst_story['H{}'.format(8+i)].value = row["reach"]
            top_worst_story['I{}'.format(8+i)].value = row["impression"]
            for letter in ['A', 'B' 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                top_worst_story['{}{}'.format(letter, 8+i)].alignment = Alignment(vertical="center")

            img = Image("img/stories.png")
            top_worst_story.add_image(img, 'C{}'.format(8+i))
        
        # Comments
        top_worst_story.merge_cells('A13:A15')
        cmt = top_worst_story['A13']
        cmt.value = 'Comments'
        top_worst_story.merge_cells('B13:I15')
        self.set_border(top_worst_story, 'B13:I15')
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="003366FF")