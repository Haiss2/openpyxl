import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class PostList():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        post_list = wb.create_sheet("Posts List")
        post_list.sheet_view.zoomScale = 85
        
        for x in list(string.ascii_uppercase[3:12]):
            post_list.column_dimensions[x].width = 14
        post_list.column_dimensions['A'].width = 24
        post_list.column_dimensions['B'].width = 24
        post_list.column_dimensions['C'].width = 14.3
        post_list.column_dimensions['D'].width = 30
        
        for merge_cells in ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:J1','K1:K2', "L1:L2"]:
            post_list.merge_cells(merge_cells)

        data = [{
                "time": "2021/05/04 18:44:09",
                "type": "Carousel",
                "thumnail": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQlu-5sicFuHii8BAVf-lwWzS0D4bOJ00mHAQ&usqp=CAU",
                "post_contents": "Styles are used to change the look of your data while displayed on screen. They are also used to determine the formatting for numbers.",
                "reach": 445,
                "impression": 500,
                "like": 23,
                "share": 45,
                "comment": 34,
                "total": 55,
                "engagement_rate": 7.48,
                "wideo_Watch_view": '-'
            }
            for i in range(20)]

        for y in range(3, 3 + len(data)):
            post_list.row_dimensions[y].height = 75
        post_list.row_dimensions[1].height = 30
        post_list.row_dimensions[2].height = 30


        for row in post_list['A1:L{}'.format(len(data)+2)]:
            for cell in row:
                cell.border = self.border

        headers = [
            { "label": "Time", "pos": "A1"},
            { "label": "Type", "pos": "B1"},
            { "label": "Thumbnail", "pos": "C1"},
            { "label": "Posts Contents", "pos": "D1"},
            { "label": "Reach", "pos": "E1"},
            { "label": "Impression", "pos": "F1"},
            { "label": "Engagement", "pos": "G1"},
            { "label": "Likes", "pos": "G2"},
            { "label": "Comments", "pos": "H2"},
            { "label": "Shares", "pos": "I2"},
            { "label": "Total", "pos": "J2"},
            { "label": "Engagement Rate", "pos": "K1"},
            { "label": "Video Watch View", "pos": "L1"},
        ]

        for header in headers:
            cell = post_list[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header["label"]
        
        for i, row in enumerate(data):
            post_list['A{}'.format(3+i)].value = row["time"]
            post_list['B{}'.format(3+i)].value = row["type"]
            post_list['D{}'.format(3+i)].value = row["post_contents"]
            post_list['E{}'.format(3+i)].value = row["reach"]
            post_list['F{}'.format(3+i)].value = row["impression"]
            post_list['G{}'.format(3+i)].value = row["like"]
            post_list['H{}'.format(3+i)].value = row["share"]
            post_list['I{}'.format(3+i)].value = row["comment"]
            post_list['J{}'.format(3+i)].value = row["total"]
            post_list['K{}'.format(3+i)].value = row["engagement_rate"]
            post_list['L{}'.format(3+i)].value = row["wideo_Watch_view"]
            for letter in ['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                post_list['{}{}'.format(letter, 3+i)].alignment = Alignment(vertical="center")
            post_list['{}{}'.format('L', 3+i)].alignment = Alignment(vertical="center", horizontal="right")

            img = Image("img/post.png")
            post_list.add_image(img, 'C{}'.format(3+i))
        
        avg_data = [662, 749, 45, 0, 20, 65, 10.14, '-']
        avg_row = len(data) + 3
         
        el = post_list['D' + str(avg_row)]
        el.value  = 'Average'
        el.font  = Font(b=True, size=12)
        el.fill = PatternFill("solid", fgColor="00969696")
        el.alignment = Alignment(horizontal="right", vertical="center")
        for i, datum in enumerate(list(string.ascii_uppercase[4:12])):
            el = post_list[datum + str(avg_row)]
            el.value = avg_data[i]
            el.font  = Font(b=True, size=12)
        
        post_list['L' + str(avg_row)].alignment = Alignment(horizontal="right", vertical="center")
        # avg border
        for row in post_list['D{}:L{}'.format(avg_row, avg_row)]:
            for cell in row:
                cell.border = self.border
    

        post_list.merge_cells('A{}:A{}'.format(len(data) + 5, len(data) + 7))
        cmt = post_list['A{}'.format(len(data) + 5)]
        cmt.value = 'Comments'
        post_list.merge_cells('B{}:L{}'.format(len(data) + 5, len(data) + 7))
        self.set_border(post_list, 'B{}:L{}'.format(len(data) + 5, len(data) + 7))
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="003366FF")

