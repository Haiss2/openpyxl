import string
import datetime as dt
import random

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class MentionRanking():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)

    def render_grid(self, ws, title, col_index, chart_where):
        # chart here
        c1 = LineChart()
        c1.width = 23.336
        c1.height = 11.2
        c1.style = 13
        c1.title = title

        labels = Reference(ws, min_row=2, max_row=32,min_col=1)
        data = Reference(ws, min_row=1, max_row=32,min_col=col_index)

        c1.add_data(data, titles_from_data=True)
        c1.set_categories(labels)

        ws.add_chart(c1, chart_where)

    def render(self):
        wb = self.wb
        mention_ranking = wb.create_sheet("Mention Ranking")
        mention_ranking.sheet_view.zoomScale = 85
        
        mention_ranking.column_dimensions['B'].width = 6
        mention_ranking.column_dimensions['C'].width = 13
        mention_ranking.column_dimensions['D'].width = 15
        mention_ranking.column_dimensions['E'].width = 4

        for i in list(string.ascii_uppercase[5:12]):
            mention_ranking.column_dimensions[i].width = 18
        
        mention_ranking.row_dimensions[1].height = 27
        for y in range(2, 60):
            mention_ranking.row_dimensions[y].height = 18.75    
        
        # merge cell here
        for merge_cells in ['A1:B1']:
            mention_ranking.merge_cells(merge_cells)

        
        mention_ranking['A1'].value = '2021/8'
        mention_ranking['A1'].alignment = Alignment(horizontal="center", vertical="center")

        mention_ranking['C1'].value = 'Number of Mention'
        mention_ranking['C1'].alignment = Alignment(horizontal="center", vertical="center")
        mention_ranking['C1'].font  = Font(b=True, color="00FFFFFF", size=9)
        mention_ranking['C1'].fill = PatternFill("solid", fgColor="003366FF")

        mention_ranking['D1'].value = 'Follower'
        mention_ranking['D1'].alignment = Alignment(horizontal="center", vertical="center")
        mention_ranking['D1'].font  = Font(b=True, color="00FFFFFF", size=9)
        mention_ranking['D1'].fill = PatternFill("solid", fgColor="003366FF")
        
        
        
        day = ['Sat','Sun','Mon','Tue','Wed','Thu','Fri']

        data = [ [str(i+1) + ' 日', day[i % 7], 0 , random.randint(1000, 1200)] for i in range(31)]

        for row in mention_ranking['A1:D{}'.format(len(data) + 2)]:
            for cell in row:
                cell.border = self.border

        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                el = mention_ranking[string.ascii_uppercase[j] + str(2+i)]
                el.value = cell
                el.number_format = 'General'
                el.alignment = Alignment(horizontal="right", vertical="center")

        self.render_grid(mention_ranking, 'Changes on Mention Number', 4, 'F2')

        # merge cell here
        for merge_cells in ['F21:L21', 'F22:F23', 'G22:H23', 'I22:I23', 'J22:L22', 'A38:B40', 'C38:L40']:
            mention_ranking.merge_cells(merge_cells)
        
        for row in mention_ranking['F22:L33']:
            for cell in row:
                cell.border = self.border

        mention_ranking['F21'].value = 'Mentioned User Ranking'
        mention_ranking['F21'].font = Font(b=False, color="00000000", size=14)
        headers = [
            { "label": "User Name", "pos": "G22"},
            { "label": "Number of Mention", "pos": "I22"},
            { "label": "Engagement", "pos": "J22"},
            { "label": "Avg. Likes", "pos": "J23"},
            { "label": "Avg. Comments", "pos": "K23"},
            { "label": "Total", "pos": "L23"},
        ]

        for header in headers:
            cell = mention_ranking[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=False, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header["label"]
        
        for i in range(1, 11):
            cell = mention_ranking['F{}'.format(23+i)]
            cell.value = 'No.' + str(i)
            cell.fill = PatternFill("solid", fgColor="00C0C0C0")

            cell = mention_ranking['L{}'.format(23+i)]
            cell.value = 0


        mention_ranking['F35'].value = '※1 The engagement figure above is the average figure of the post being tagged/mentioned by each user. '
        mention_ranking['F36'].value = '※2 Mention data can only be available after the day your acc was connected to Reposta. '

        mention_ranking['A38'].value = 'Comments'
        self.set_border(mention_ranking, 'C38:L40')
        mention_ranking['A38'].alignment = Alignment(horizontal="center", vertical="center")
        mention_ranking['A38'].font  = Font(b=True, color="00FFFFFF", size=12)
        mention_ranking['A38'].fill = PatternFill("solid", fgColor="003366FF")

        