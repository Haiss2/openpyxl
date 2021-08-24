import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import PieChart, ProjectedPieChart, Reference
from openpyxl.chart.series import DataPoint

class CoverPage():

    def __init__(self, wb):
        self.wb = wb
    
    def render(self):
        wb = self.wb
        cover_page = wb.create_sheet("Cover Page", 0) 
        cover_page.sheet_view.showGridLines = False

        cover_page.row_dimensions[11].height = 25
        cover_page.row_dimensions[12].height = 25
        cover_page.row_dimensions[13].height = 30
        cover_page.row_dimensions[15].height = 25

        for merge_cells in ['A11:N11', 'A12:N12', 'A13:N13', 'A15:N15']:
            cover_page.merge_cells(merge_cells)

        title = cover_page['A11']
        title.value = "Instagram Analytic Report"
        title.fill = PatternFill("solid", fgColor="003366FF")
        title.font  = Font(color="00FFFFFF")
        title.alignment = Alignment(horizontal="center", vertical="center")

        account  = cover_page['A12']
        account.value = "Account Name"
        account.font  = Font(color="00808080")
        account.alignment = Alignment(horizontal="center", vertical="center")

        user = cover_page['A13']
        user.value = "iwai_kampo"
        user.font  = Font(color="00000000", size=13)
        user.alignment = Alignment(horizontal="center", vertical="center")


        user = cover_page['A15']
        user.value = dt.datetime.now().strftime("%b/%d/%Y")
        user.font  = Font(color="00000000", size=10)
        user.alignment = Alignment(horizontal="center", vertical="center")