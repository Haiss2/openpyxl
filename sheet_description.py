import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class SheetDescription():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)
    
    def render(self):
        wb = self.wb
        sheet_des = wb.create_sheet("Sheet Description") 
        sheet_des.sheet_view.zoomScale = 120

        sheet_des.column_dimensions['A'].width = 45
        sheet_des.column_dimensions['B'].width = 80

        for row in sheet_des['A1:B17']:
            for cell in row:
                cell.border = self.border

    
        sheet_des['A1'].value = "Name of Sheet"
        sheet_des['A1'].fill = PatternFill("solid", fgColor="003366FF")
        sheet_des['A1'].font  = Font(b=True, color="00FFFFFF")
        sheet_des['A1'].alignment = Alignment(horizontal="center", vertical="center")

        sheet_des['B1'].value = "Description"
        sheet_des['B1'].fill = PatternFill("solid", fgColor="003366FF")
        sheet_des['B1'].font  = Font(b=True, color="00FFFFFF")
        sheet_des['B1'].alignment = Alignment(horizontal="center", vertical="center")

        data = [
            ("Cover Page", "Cover page of this report", 18),
            ("Follower Analysis", "フォロワーの属性を確認することができます。", 18),
            ("MoM", "Month to Month change for past 6 months from the month you pointed.", 18),
            ("DoD", "You can check changes on daily basis for Number of follower, increase & decrease \n of follower(+/-), Profile View, Website clicks, Reach, Impression.", 35),
            ("Posts List", "You can check detail data of each posts.", 18),
            ("Ads Post List\n※For plan above Plus", "You can check the post you made trhough Ads Manager.*", 35),
            ("Stories Post List", "You can check detail data of each stories.", 18),
            ("Avg. of each type of Post", "You can check the Avg. number of each type of post.", 18),
            ("Post - Engagement Top&Worst3", "You can check Top3 & Worst3 Engament number of post in Ranking format.", 18),
            ("Stories - Impression TOP&WORST3", "You can check Top3 & Worst3 Impression of Stories in Ranking format.", 35),
            ("Mention Ranking\n※Only for plan above Basic", "メンション・タグ付けをされたユーザーのランキングとそのメンション\n数を確認することができます。", 35),
            ("Competitor Comparison\n※Only for plan above Basic", "You can check number of follower and post dates of your competitors by comparing to \n your account.", 35),
            ("Competitor Engagement TOP&WORST3\n※Only for plan above Basic", "You can check Top3 & Worst3 Engagement number of post of competitor.", 35),
            ("Hushtag Analysis\n※Only for plan above Plus", "You can check the related hushtag and if the hushtag is used on the popular post. You can \n analyize if the hushtag is effective or not.\n\n※Popular posts are optimized for each user, therefore this report result might be \ndifferent from the one you see on your app.", 88),
            ("Ranked in Hushtag\n※Only for plan above Basic", "You can check the detail of the post that is listed on Popular post.\n\n※Popular posts are optimized for each user, therefore this report result might be \n different from the one you see on your app.", 72),
            ("Word Description",	"Description of words used in this report.", 18)
        ] 
        
        for i, datum in enumerate(data):
            sheet_des['A' + str(i+2)].value = datum[0]
            sheet_des['A' + str(i+2)].alignment = Alignment(vertical="center", wrap_text=True)

            sheet_des['B' + str(i+2)].value = datum[1]
            sheet_des['B' + str(i+2)].alignment = Alignment(vertical="center", wrap_text=True)

            sheet_des.row_dimensions[i+2].height = datum[2]


        sheet_des['A19'].value = "※The figure of advertisement can only be available on DoD and Ads Post List."











