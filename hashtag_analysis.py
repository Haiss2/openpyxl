import string
import datetime as dt
import random

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class HashtagAnalysis():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)

    def render_grid(self, ws, title, col_index, chart_where):
        # chart here
        c1 = LineChart()
        c1.width = 23.336
        c1.height = 11.2
        c1.style = 13
        c1.title = title

        labels = Reference(ws, min_row=3, max_row=33,min_col=1)
        data = Reference(ws, min_row=2, max_row=33,min_col=col_index)

        c1.add_data(data, titles_from_data=True)
        c1.set_categories(labels)

        ws.add_chart(c1, chart_where)

    def render(self):
        wb = self.wb
        hashtag_analysis = wb.create_sheet("Hashtag Analysis")
        hashtag_analysis.sheet_view.zoomScale = 85
        
        hashtag_analysis.column_dimensions['B'].width = 6
        hashtag_analysis.column_dimensions['C'].width = 13
        hashtag_analysis.column_dimensions['D'].width = 15
        hashtag_analysis.column_dimensions['E'].width = 15
        hashtag_analysis.column_dimensions['F'].width = 4

        for i in list(string.ascii_uppercase[6:13]):
            hashtag_analysis.column_dimensions[i].width = 18
        
        hashtag_analysis.row_dimensions[1].height = 27
        hashtag_analysis.row_dimensions[1].height = 27
        for y in range(3, 60):
            hashtag_analysis.row_dimensions[y].height = 18.75    
        
        # merge cell here
        for merge_cells in ['A1:E1', 'A2:B2']:
            hashtag_analysis.merge_cells(merge_cells)

        
        hashtag_analysis['A1'].value = '#非結核性抗酸菌症'
        hashtag_analysis['A1'].alignment = Alignment(horizontal="center", vertical="center")

        hashtag_analysis['C2'].value = 'Number of Ranked in'
        hashtag_analysis['C2'].alignment = Alignment(horizontal="center", vertical="center")
        hashtag_analysis['C2'].font  = Font(b=True, color="00FFFFFF", size=9)
        hashtag_analysis['C2'].fill = PatternFill("solid", fgColor="003366FF")

        hashtag_analysis['D2'].value = 'Follower'
        hashtag_analysis['D2'].alignment = Alignment(horizontal="center", vertical="center")
        hashtag_analysis['D2'].font  = Font(b=True, color="00FFFFFF", size=9)
        hashtag_analysis['D2'].fill = PatternFill("solid", fgColor="003366FF")

        hashtag_analysis['E2'].value = '+/-'
        hashtag_analysis['E2'].alignment = Alignment(horizontal="center", vertical="center")
        hashtag_analysis['E2'].font  = Font(b=True, color="00FFFFFF", size=9)
        hashtag_analysis['E2'].fill = PatternFill("solid", fgColor="003366FF")
        
        
        
        day = ['Sat','Sun','Mon','Tue','Wed','Thu','Fri']

        data = [ [str(i+1) + ' 日', day[i % 7], 0 , random.randint(1000, 1200)] for i in range(31)]

        for row in hashtag_analysis['A2:E{}'.format(len(data) + 3)]:
            for cell in row:
                cell.border = self.border

        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                el = hashtag_analysis[string.ascii_uppercase[j] + str(3+i)]
                el.value = cell
                el.number_format = 'General'
                el.alignment = Alignment(horizontal="right", vertical="center")
            Eel = hashtag_analysis['E'+ str(3+i)]
            if i == 0:
               Eel.value = '-' 
            else:
                Eel.value = data[i][3] - data[i-1][3]

        self.render_grid(hashtag_analysis, 'Changes on Number of Ranked in & Followers', 4, 'G2')

        # merge cell here
        for merge_cells in ['G21:M21', 'G22:G23', 'H22:I23', 'J22:J23', 'K22:M22', 'A40:B42', 'C40:M42']:
            hashtag_analysis.merge_cells(merge_cells)
        
        for row in hashtag_analysis['G22:M33']:
            for cell in row:
                cell.border = self.border

        hashtag_analysis['G21'].value = 'Related Hushtags'
        hashtag_analysis['G21'].font = Font(b=False, color="00000000", size=14)
        headers = [
            { "label": "Hashtag", "pos": "H22"},
            { "label": "User Rate", "pos": "J22"},
            { "label": "Engagement", "pos": "K22"},
            { "label": "Avg. Likes", "pos": "K23"},
            { "label": "Avg. Comments", "pos": "L23"},
            { "label": "Total", "pos": "M23"},
        ]

        for header in headers:
            cell = hashtag_analysis[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=False, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header["label"]
        
        for i in range(1, 11):
            cell = hashtag_analysis['G{}'.format(23+i)]
            cell.value = 'No.' + str(i)
            cell.fill = PatternFill("solid", fgColor="00C0C0C0")

            cell = hashtag_analysis['M{}'.format(23+i)]
            cell.value = 0


        hashtag_analysis['F36'].value = '※1　The extracted Related hushtag is most commonly used among popular posts.'
        hashtag_analysis['F37'].value = '※2　Used Rate is the rate how often the hushtag is used among popular posts.'
        hashtag_analysis['F38'].value = '※3　The ranked in Number is the number your post is ranked in within top100 among pupular posts.'

        hashtag_analysis['A40'].value = 'Comments'
        self.set_border(hashtag_analysis, 'C40:M42')
        hashtag_analysis['A40'].alignment = Alignment(horizontal="center", vertical="center")
        hashtag_analysis['A40'].font  = Font(b=True, color="00FFFFFF", size=12)
        hashtag_analysis['A40'].fill = PatternFill("solid", fgColor="003366FF")

        