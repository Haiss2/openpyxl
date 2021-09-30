import string
import datetime as dt
import random

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class CompetitorComparision():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)

    def render_grid(self, ws, title, col_label_index, col_data_index, chart_where, row_start, row_end, width, height):
        # chart here
        c1 = LineChart()
        c1.width = width
        c1.height = height
        c1.style = 13
        c1.title = title

        labels = Reference(ws, min_row=row_start, max_row=row_end,min_col=col_label_index)
        data = Reference(ws, min_row=row_start-1, max_row=row_end,min_col=col_data_index)

        c1.add_data(data, titles_from_data=True)
        c1.set_categories(labels)

        ws.add_chart(c1, chart_where)

    def render(self):
        wb = self.wb
        competitor_comparision = wb.create_sheet("Competitor Comparison")
        competitor_comparision.sheet_view.zoomScale = 85
        competitor_comparision.column_dimensions['A'].width = 6
        competitor_comparision.column_dimensions['B'].width = 6

        for i in list(string.ascii_uppercase[2:8]):
            competitor_comparision.column_dimensions[i].width = 13

        competitor_comparision.column_dimensions['I'].width = 4

        for i in list(string.ascii_uppercase[9:19]):
            competitor_comparision.column_dimensions[i].width = 13
        
        competitor_comparision.row_dimensions[1].height = 20
        competitor_comparision.row_dimensions[2].height = 27
        for y in range(3, 60):
            competitor_comparision.row_dimensions[y].height = 18.75    
        
        # merge cell here
        for merge_cells in ['A1:B2', 'C1:E1', 'F1:H1']:
            competitor_comparision.merge_cells(merge_cells)
        
        _1st_table_headers = [
            {"value": "airmega_coway", "pos": 'C1', "fill_color": "00FFCC99", "color": "00000000", "b": False},
            {"value": "Post", "pos": 'C2', "fill_color": "00FFCC99", "color": "00000000", "b": False},
            {"value": "Follower", "pos": 'D2', "fill_color": "00FFCC99", "color": "00000000", "b": False},
            {"value": "Follower\n(+/-)", "pos": 'E2', "fill_color": "00FFCC99", "color": "00000000", "b": False},
            
            {"value": "iwai_kampo", "pos": 'F1', "fill_color": "003366FF", "color": "00FFFFFF", "b": True},
            {"value": "Post", "pos": 'F2', "fill_color": "003366FF", "color": "00FFFFFF", "b": True},
            {"value": "Follower", "pos": 'G2', "fill_color": "003366FF", "color": "00FFFFFF", "b": True},
            {"value": "Follower\n(+/-)", "pos": 'H2', "fill_color": "003366FF", "color": "00FFFFFF", "b": True}
        ]
        
        for header in _1st_table_headers:
            el = competitor_comparision[header["pos"]]
            el.value = header["value"]
            el.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            el.font  = Font(b=True, color=header["color"], size=9)
            el.fill = PatternFill("solid", fgColor=header["fill_color"])

        
        day = ['Sat','Sun','Mon','Tue','Wed','Thu','Fri']

        days = [ [str(i+1) + ' 日', day[i % 7]] for i in range(31)]
        competitive_data = [ [random.randint(0, 3), random.randint(1000, 1200)] for i in range(31)]
        my_data = [ [random.randint(0, 3) , random.randint(1000, 1200)] for i in range(31)]

        for row in competitor_comparision['A1:H{}'.format(len(days) + 3)]:
            for cell in row:
                cell.border = self.border

        for i in range(len(days)):
            for j in range(8):
                el = competitor_comparision[string.ascii_uppercase[j] + str(3+i)]
                if j == 0 or j == 1:
                    el.value = days[i][j]
                if j == 2 or j == 3:
                    el.value = competitive_data[i][j-2]
                if j == 5 or j == 6:
                    el.value = my_data[i][j-5]
                el.number_format = 'General'
                el.alignment = Alignment(horizontal="right", vertical="center")
            Eel = competitor_comparision['E'+ str(3+i)]
            Hel = competitor_comparision['H'+ str(3+i)]
            Eel.value = '-' if i == 0 else competitive_data[i][1] - competitive_data[i-1][1]
            Hel.value = '-' if i == 0 else my_data[i][1] - my_data[i-1][1]

        self.render_grid(competitor_comparision, 'Number of Followers',1, 4, 'J1', 3, 33, 24, 7.53)
        self.render_grid(competitor_comparision, 'Number of Posts', 1, 3, 'J13', 3, 33, 24, 7)

        # merge cell here
        for merge_cells in ['J25:L25', 'J26:J27', 'K26:K27', 'L26:L27', 'A39:B41', 'C39:S41']:
            competitor_comparision.merge_cells(merge_cells)
        
        for row in competitor_comparision['J26:L34']:
            for cell in row:
                cell.border = self.border

        competitor_comparision['J25'].value = 'Number of Post on Date'
        competitor_comparision['J25'].font = Font(b=False, color="00000000", size=14)
 

        
        competitor_comparision['K26'].alignment = Alignment(horizontal="center", vertical="center")
        competitor_comparision['K26'].font  = Font(b=False, color="00000000", size=12)
        competitor_comparision['K26'].fill = PatternFill("solid", fgColor="00FFCC99")
        competitor_comparision['K26'].value = "airmega_coway"


        competitor_comparision['L26'].alignment = Alignment(horizontal="center", vertical="center")
        competitor_comparision['L26'].font  = Font(b=False, color="00FFFFFF", size=12)
        competitor_comparision['L26'].fill = PatternFill("solid", fgColor="003366FF")
        competitor_comparision['L26'].value = "iwai_kampo"
            
        for i, date in enumerate(day):
            cell = competitor_comparision['J{}'.format(28+i)]
            cell.value = date
            cell.fill = PatternFill("solid", fgColor="00C0C0C0")
            
            competitor_comparision['K{}'.format(28+i)].value = random.randint(0,6) 
            competitor_comparision['L{}'.format(28+i)].value = random.randint(0,6)
        
        self.render_grid(competitor_comparision, 'Number of Post on Date', 10, 11, 'N25', 28, 34, 14.4, 6.5)


        competitor_comparision['J36'].value = '※1 The follower of competitor is available since the day you connected your account to Reposta.'
        competitor_comparision['J37'].value = '※2 The follower of account is available since the day you connected your account to Reposta.'

        competitor_comparision['A39'].value = 'Comments'
        self.set_border(competitor_comparision, 'C39:S41')
        competitor_comparision['A39'].alignment = Alignment(horizontal="center", vertical="center")
        competitor_comparision['A39'].font  = Font(b=True, color="00FFFFFF", size=12)
        competitor_comparision['A39'].fill = PatternFill("solid", fgColor="003366FF")

        