import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class CompetitorTopWorst():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        competitor_top_worst = wb.create_sheet("Competitor Engagemen TOP&WORST3")
        competitor_top_worst.sheet_view.zoomScale = 85
        
        for x in list(string.ascii_uppercase[4:12]):
            competitor_top_worst.column_dimensions[x].width = 14
        competitor_top_worst.column_dimensions['A'].width = 14
        competitor_top_worst.column_dimensions['B'].width = 24
        competitor_top_worst.column_dimensions['C'].width = 24
        competitor_top_worst.column_dimensions['D'].width = 14.3
        competitor_top_worst.column_dimensions['E'].width = 30
        
        for merge_cells in ['A1:B1','A2:A3', 'B2:B3', 'C2:C3', 'D2:D3', 'E2:E3', 'F2:H2']:
            competitor_top_worst.merge_cells(merge_cells)

        headers = [
            { "label": "", "pos": "A2"},
            { "label": "Posted Date", "pos": "B2"},
            { "label": "Type", "pos": "C2"},
            { "label": "Thumbnail", "pos": "D2"},
            { "label": "Posts Contents", "pos": "E2"},
            { "label": "Engagement", "pos": "F2"},
            { "label": "Likes", "pos": "F3"},
            { "label": "Comments", "pos": "G3"},
            { "label": "Total", "pos": "H3"},
        ]

        for header in headers:
            cell = competitor_top_worst[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="00FFCC99")
            cell.value = header["label"]


        datum = {
                "time": "2021/05/04 18:44:09",
                "type": "Carousel",
                "thumnail": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQlu-5sicFuHii8BAVf-lwWzS0D4bOJ00mHAQ&usqp=CAU",
                "post_contents": "Styles are used to change the look of your data while displayed on screen. They are also used to determine the formatting for numbers.",
                "like": 23,
                "comment": 34,
                "total": 55,
        }

        competitor_top_worst.row_dimensions[1].height = 30
        competitor_top_worst.row_dimensions[2].height = 15
        competitor_top_worst.row_dimensions[2].height = 15
        competitor_top_worst.row_dimensions[4].height = 15
        competitor_top_worst.row_dimensions[8].height = 15

        top_data = [ datum for i in range(3)]
        worst_data = [ datum for i in range(3)]

        for y in [5,6,7,9,10, 11]:
            competitor_top_worst.row_dimensions[y].height = 70

        for row in competitor_top_worst['A2:H11']:
            for cell in row:
                cell.border = self.border
        
        # Top of the world
        competitor_top_worst['A4'].value = 'TOP▲'
        competitor_top_worst['A4'].font = Font(b=True, color="00000000", size=12)
        for row in competitor_top_worst['A4:H4']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="0099CC00")
        
        for i, row in enumerate(top_data):
            no = competitor_top_worst['A{}'.format(5+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00000000", size=16)
            no.fill = PatternFill("solid", fgColor="0099CC00")

            competitor_top_worst['B{}'.format(5+i)].value = row["time"]
            competitor_top_worst['C{}'.format(5+i)].value = row["type"]
            competitor_top_worst['E{}'.format(5+i)].value = row["post_contents"]
            competitor_top_worst['F{}'.format(5+i)].value = row["like"]
            competitor_top_worst['G{}'.format(5+i)].value = row["comment"]
            competitor_top_worst['H{}'.format(5+i)].value = row["total"]
            for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                competitor_top_worst['{}{}'.format(letter, 5+i)].alignment = Alignment(vertical="center")

            img = Image("img/post.png")
            competitor_top_worst.add_image(img, 'D{}'.format(5+i))
        

        # Worst of the world
        competitor_top_worst['A8'].value = 'WORST▼'
        competitor_top_worst['A8'].font = Font(b=True, color="00FFFFFF", size=12)
        for row in competitor_top_worst['A8:H8']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="00FF6600")
        
        for i, row in enumerate(worst_data):
            no = competitor_top_worst['A{}'.format(8+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00FFFFFF", size=16)
            no.fill = PatternFill("solid", fgColor="00FF6600")

            competitor_top_worst['B{}'.format(9+i)].value = row["time"]
            competitor_top_worst['C{}'.format(9+i)].value = row["type"]
            competitor_top_worst['E{}'.format(9+i)].value = row["post_contents"]
            competitor_top_worst['F{}'.format(9+i)].value = row["like"]
            competitor_top_worst['G{}'.format(9+i)].value = row["comment"]
            competitor_top_worst['H{}'.format(9+i)].value = row["total"]
            for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                competitor_top_worst['{}{}'.format(letter, 9+i)].alignment = Alignment(vertical="center")

            img = Image("img/post.png")
            competitor_top_worst.add_image(img, 'D{}'.format(9+i))
        

        # Notions
        competitor_top_worst['F13'].value = "※1Thumbnail of movie won't be displayed."

        # Comments
        competitor_top_worst.merge_cells('A15:B17')
        cmt = competitor_top_worst['A15']
        cmt.value = 'Comments'
        competitor_top_worst.merge_cells('C15:H17')
        self.set_border(competitor_top_worst, 'C15:H17')
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="00FFCC99")
       

