import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class DoD():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)

    def render_grid(self, ws, title, col_index, chart_where):
        # chart here
        c1 = LineChart()
        c1.width = 18
        c1.height = 9
        c1.style = 13
        c1.title = title

        labels = Reference(ws, min_row=3, max_row=35,min_col=1)
        data = Reference(ws, min_row=2, max_row=35,min_col=col_index)

        c1.add_data(data, titles_from_data=True)
        c1.set_categories(labels)

        ws.add_chart(c1, chart_where)

    def render(self):
        wb = self.wb
        d_o_d = wb.create_sheet("DoD")
        d_o_d.sheet_view.zoomScale = 70
        
        for x in list(string.ascii_uppercase[2:]):
            d_o_d.column_dimensions[x].width = 16
        for y in range(3, 60):
            d_o_d.row_dimensions[y].height = 27
        d_o_d.row_dimensions[1].height = 45
        d_o_d.row_dimensions[2].height = 45
        
        for merge_cells in ['A1:H1', 'A39:C41', 'D39:V41']:
            d_o_d.merge_cells(merge_cells)
        
        d_o_d['A1'].value = '2021/8'
        d_o_d['A1'].font  = Font(b=True, size=18)
        d_o_d['A1'].alignment = Alignment(horizontal="left", vertical="center")

        data = [
            ['1 日', 'Sat', 1078 ,'-', 5 , 0, 114, 131],
            ['2 日', 'Sun', 1078 ,0 , 4, 0, 26, 31 ],
            ['3 日', 'Mon', 1078 ,0 , 0, 0, 18,	19 ],
            ['4 日', 'Tue', 1078 ,0 , 7, 0, 144, 164 ],
            ['5 日', 'Wed', 1078 ,1, 12, 0, 41, 58 ],
            ['6 日', 'Thu', 1078 ,1 , 1, 0, 31, 33 ],
            ['7 日', 'Fri', 1078 ,2 , 11, 0, 139, 179 ],
        ]
        data = [*data,*data,*data,*data,['29 日', 'Thu', '1,078','1' , '1', '0', '31', '33' ],['30 日', 'Fri', '1,080','2' , '11', '0', '139', '179' ], ['31 日', 'Fri', '1,080','2' , '11', '0', '139', '179' ]]
        headers = ['Day', 'Date', 'Follower', 'Follower\n(+/-)', 'Profile View', 'Website\nClicks', 'Reach', 'Impression']

        self.set_border(d_o_d, 'A3:H33')

        for index, header in enumerate(headers):
            element = d_o_d[string.ascii_uppercase[index] + str(2)]
            element.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            element.value = header
            element.font  = Font(b=True, color="00FFFFFF")
            element.fill = PatternFill("solid", fgColor="003366FF")

        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                el = d_o_d[string.ascii_uppercase[j] + str(3+i)]
                el.value = cell
                el.number_format = 'General'
                el.alignment = Alignment(horizontal="right", vertical="center")

        self.render_grid(d_o_d, 'Follower', 3, 'J2')
        self.render_grid(d_o_d, 'Follower (+/-)', 4, 'Q2')
        self.render_grid(d_o_d, 'Profile View', '5', 'J12')
        self.render_grid(d_o_d, 'Website Clicks', 6, 'Q12')
        self.render_grid(d_o_d, 'Reach', 7, 'J23')
        self.render_grid(d_o_d, 'Impression', 8, 'Q23')

        d_o_d['K34'].value = '※1 The figure of Follower & Follower (+/-) can only be shown since the day your acc is connected with Reposta'
        d_o_d['K35'].value = '※2 The figure includes the advertisement as well.'

        d_o_d['A39'].value = 'Comments'
        self.set_border(d_o_d, 'D39:V41')
        d_o_d['A39'].alignment = Alignment(horizontal="center", vertical="center")
        d_o_d['A39'].font  = Font(b=True, color="00FFFFFF", size=12)
        d_o_d['A39'].fill = PatternFill("solid", fgColor="003366FF")