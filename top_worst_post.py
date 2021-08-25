import string
import datetime as dt

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis

class TopWorstPost():

    def __init__(self, wb):
        self.wb = wb
        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            row[0].border = Border(left=self.thin)
            row[-1].border = Border(right=self.thin)
        for c in rows[0]:
            c.border = Border(top=self.thin)
        for c in rows[-1]:
            c.border = Border(bottom=self.thin)
        rows[0][0].border = Border(top=self.thin, left=self.thin)
        rows[0][-1].border = Border(top=self.thin, right=self.thin)
        rows[-1][0].border = Border(bottom=self.thin, left=self.thin)
        rows[-1][-1].border = Border(bottom=self.thin, right=self.thin)


    def render(self):
        wb = self.wb
        top_worst_post = wb.create_sheet("Post-Engagement TOP&WORST3")
        top_worst_post.sheet_view.zoomScale = 85
        
        for x in list(string.ascii_uppercase[4:12]):
            top_worst_post.column_dimensions[x].width = 14
        top_worst_post.column_dimensions['A'].width = 14
        top_worst_post.column_dimensions['B'].width = 24
        top_worst_post.column_dimensions['C'].width = 24
        top_worst_post.column_dimensions['D'].width = 14.3
        top_worst_post.column_dimensions['E'].width = 30
        
        for merge_cells in ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:K1', "L1:L2"]:
            top_worst_post.merge_cells(merge_cells)

        headers = [
            { "label": "", "pos": "A1"},
            { "label": "Time", "pos": "B1"},
            { "label": "Type", "pos": "C1"},
            { "label": "Thumbnail", "pos": "D1"},
            { "label": "Posts Contents", "pos": "E1"},
            { "label": "Reach", "pos": "F1"},
            { "label": "Impression", "pos": "G1"},
            { "label": "Engagement", "pos": "H1"},
            { "label": "Likes", "pos": "H2"},
            { "label": "Comments", "pos": "I2"},
            { "label": "Shares", "pos": "J2"},
            { "label": "Total", "pos": "K2"},
            { "label": "Engagement Rate", "pos": "L1"},
        ]

        for header in headers:
            cell = top_worst_post[header["pos"]]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font  = Font(b=True, color="00FFFFFF", size=12)
            cell.fill = PatternFill("solid", fgColor="003366FF")
            cell.value = header["label"]



        datum = {
                "time": "2021/05/04 18:44:09",
                "type": "Carousel",
                "thumnail": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQlu-5sicFuHii8BAVf-lwWzS0D4bOJ00mHAQ&usqp=CAU",
                "post_contents": "Styles are used to change the look of your data while displayed on screen. They are also used to determine the formatting for numbers.",
                "reach": 445,
                "impression": 500,
                "like": 23,
                "share": 45,
                "comment": 34,
                "total": 55,
                "engagement_rate": 7.48,
        }
        top_worst_post.row_dimensions[1].height = 30
        top_worst_post.row_dimensions[2].height = 30

        top_data = [ datum for i in range(3)]
        worst_data = [ datum for i in range(3)]

        for y in [4,5,6,8,9,10]:
            top_worst_post.row_dimensions[y].height = 75


        for row in top_worst_post['A1:L10']:
            for cell in row:
                cell.border = self.border
        
        # Top of the world
        top_worst_post['A3'].value = 'TOP▲'
        top_worst_post['A3'].font = Font(b=True, color="00000000", size=12)
        for row in top_worst_post['A3:L3']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="0099CC00")
        
        for i, row in enumerate(top_data):
            no = top_worst_post['A{}'.format(4+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00000000", size=16)
            no.fill = PatternFill("solid", fgColor="0099CC00")

            top_worst_post['B{}'.format(4+i)].value = row["time"]
            top_worst_post['C{}'.format(4+i)].value = row["type"]
            top_worst_post['E{}'.format(4+i)].value = row["post_contents"]
            top_worst_post['F{}'.format(4+i)].value = row["reach"]
            top_worst_post['G{}'.format(4+i)].value = row["impression"]
            top_worst_post['H{}'.format(4+i)].value = row["like"]
            top_worst_post['I{}'.format(4+i)].value = row["share"]
            top_worst_post['J{}'.format(4+i)].value = row["comment"]
            top_worst_post['K{}'.format(4+i)].value = row["total"]
            top_worst_post['L{}'.format(4+i)].value = row["engagement_rate"]
            for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                top_worst_post['{}{}'.format(letter, 4+i)].alignment = Alignment(vertical="center")

            img = Image("img/post.png")
            top_worst_post.add_image(img, 'D{}'.format(4+i))
        

        # Worst of the world
        top_worst_post['A7'].value = 'WORST▼'
        top_worst_post['A7'].font = Font(b=True, color="00FFFFFF", size=12)
        for row in top_worst_post['A7:L7']:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="00FF6600")
        
        for i, row in enumerate(worst_data):
            no = top_worst_post['A{}'.format(8+i)]
            no.value = 'No.' + str(i+1)
            no.font = Font(b=True, color="00FFFFFF", size=16)
            no.fill = PatternFill("solid", fgColor="00FF6600")

            top_worst_post['B{}'.format(8+i)].value = row["time"]
            top_worst_post['C{}'.format(8+i)].value = row["type"]
            top_worst_post['E{}'.format(8+i)].value = row["post_contents"]
            top_worst_post['F{}'.format(8+i)].value = row["reach"]
            top_worst_post['G{}'.format(8+i)].value = row["impression"]
            top_worst_post['H{}'.format(8+i)].value = row["like"]
            top_worst_post['I{}'.format(8+i)].value = row["share"]
            top_worst_post['J{}'.format(8+i)].value = row["comment"]
            top_worst_post['K{}'.format(8+i)].value = row["total"]
            top_worst_post['L{}'.format(8+i)].value = row["engagement_rate"]
            for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                top_worst_post['{}{}'.format(letter, 8+i)].alignment = Alignment(vertical="center")

            img = Image("img/post.png")
            top_worst_post.add_image(img, 'D{}'.format(8+i))
        

        # Notions
        top_worst_post['E12'].value = "※1 The recent post can be ranked in WORST as its engagement might not have been increasing much. "

        # Comments
        top_worst_post.merge_cells('A14:A16')
        cmt = top_worst_post['A14']
        cmt.value = 'Comments'
        top_worst_post.merge_cells('B14:L16')
        self.set_border(top_worst_post, 'B14:L16')
        cmt.alignment = Alignment(horizontal="center", vertical="center")
        cmt.font  = Font(b=True, color="00FFFFFF", size=12)
        cmt.fill = PatternFill("solid", fgColor="003366FF")
       

